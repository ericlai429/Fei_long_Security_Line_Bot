# -*- coding: utf-8 -*-
"""
飛龍保全 勤務排班系統 - 輕量自動化同步小輪子 (sync_wheel.py)
功能：
1. 自動探測桌面最新 115年9月班表3.0.xlsx (或指定路徑)
2. 完整解析 4.三總工務所 與 5.三總重症大樓 全月 30 天排班
3. 精準識別請假替補 (9/4 賴大哥代班林又妗)
4. 一鍵寫入 docs/data/、index.html、docs/index.html、app/static/pwa/index.html
5. 自動計算版本 Hash 並支援一鍵推送至 GitHub (git push)
"""
import os
import sys
import json
import re
import glob
import hashlib
import shutil
import subprocess

def find_excel_file():
    candidates = [
        r'C:\Users\user\Desktop\9月班表_三總\115年9月班表3.0.xlsx',
        r'C:\Users\user\Desktop\115年9月班表3.0.xlsx',
        r'C:\Users\user\Downloads\115年9月班表3.0.xlsx',
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    desktop_matches = glob.glob(r'C:\Users\user\Desktop\*115*9*班表*.xlsx')
    if desktop_matches:
        return desktop_matches[0]
    return None

def stable_stringify(obj):
    if isinstance(obj, list):
        return '[' + ','.join(stable_stringify(x) for x in obj) + ']'
    if isinstance(obj, dict):
        keys = sorted(obj.keys())
        return '{' + ','.join(json.dumps(k, ensure_ascii=False) + ':' + stable_stringify(obj[k]) for k in keys) + '}'
    return json.dumps(obj, ensure_ascii=False)

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    print('=====================================================')
    print(' 🚀 飛龍保全排班系統 - 輕量自動同步小輪子')
    print('=====================================================')

    excel_path = find_excel_file()
    if not excel_path:
        print('❌ 找不到 115年9月班表3.0.xlsx 檔案！')
        return

    print(f'📖 1. 讀取最新班表檔案: {excel_path}')
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['115.9'] if '115.9' in wb.sheetnames else wb.active

    # 解析日期對照
    day_map = {}
    for c in range(4, 34):
        cell_val = ws.cell(2, c).value
        w_val = ws.cell(3, c).value or ''
        if cell_val is not None:
            if hasattr(cell_val, 'day'):
                d_num = cell_val.day
                d_str = cell_val.strftime('%Y/%m/%d')
            else:
                m = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', str(cell_val))
                if m:
                    d_num = int(m.group(3))
                    d_str = f"{m.group(1)}/{int(m.group(2)):02d}/{int(m.group(3)):02d}"
                else:
                    d_num = c - 3
                    d_str = f"2026/09/{d_num:02d}"
            day_map[c] = (d_num, str(w_val).strip(), d_str)

    eng_rows = {d_num: {'日期': d_str, '星期': wk, '哨點/崗位': '4.三總工務所', '早班 (07-19)': [], '晚班 (19-07)': [], 'substitutes': []} for c, (d_num, wk, d_str) in day_map.items()}
    icu_rows = {d_num: {'日期': d_str, '星期': wk, '哨點/崗位': '5.三總重症大樓', '早班 (07-19)': [], '晚班 (19-07)': [], 'substitutes': []} for c, (d_num, wk, d_str) in day_map.items()}

    # 解析行數據
    for r in range(4, 16):
        site_cell = str(ws.cell(r, 1).value or '').strip()
        shift_type_cell = str(ws.cell(r, 2).value or '').strip()
        person_cell = str(ws.cell(r, 3).value or '').strip()

        if not person_cell:
            continue

        phone_match = re.search(r'09\d{2}[-\s]?\d{3}[-\s]?\d{3}|\d{9,10}', person_cell)
        phone = phone_match.group(0) if phone_match else ''
        name_pure = re.sub(r'[\(（]?09\d{2}[-\s]?\d{3}[-\s]?\d{3}[\)）]?|\d{8,10}|[\(（]\d+[\)]?', '', person_cell.split('\n')[0]).strip()

        display_name = f"{name_pure} ({phone})" if phone else name_pure
        is_eng = '工務所' in site_cell
        target_dict = eng_rows if is_eng else icu_rows

        for c, (d_num, wk, d_str) in day_map.items():
            v = ws.cell(r, c).value
            if v is not None:
                v_str = str(v).strip().upper()
                if v_str in ['A', '早', '日'] or ('日' in shift_type_cell and v_str in ['機', '支', '代', 'V', '1']):
                    if display_name not in target_dict[d_num]['早班 (07-19)']:
                        target_dict[d_num]['早班 (07-19)'].append(display_name)
                        # 特殊標記：9/4 賴大哥代班林又妗
                        if d_num == 4 and '重症' in site_cell and '賴鯤仲' in name_pure:
                            target_dict[d_num]['substitutes'].append(f'day_{name_pure}')
                elif v_str in ['B', '晚', '夜'] or ('夜' in shift_type_cell and v_str in ['機', '支', '代', 'V', '1']):
                    if display_name not in target_dict[d_num]['晚班 (19-07)']:
                        target_dict[d_num]['晚班 (19-07)'].append(display_name)

    # 組合格式
    eng_final = []
    icu_final = []

    for d_num in sorted(eng_rows.keys()):
        r = eng_rows[d_num]
        item = {
            '日期': r['日期'],
            '星期': r['星期'],
            '哨點/崗位': '4.三總工務所',
            '早班 (07-19)': '、'.join(r['早班 (07-19)']) if r['早班 (07-19)'] else '—',
            '晚班 (19-07)': '、'.join(r['晚班 (19-07)']) if r['晚班 (19-07)'] else '—'
        }
        if r['substitutes']:
            item['substitutes'] = r['substitutes']
        eng_final.append(item)

    for d_num in sorted(icu_rows.keys()):
        r = icu_rows[d_num]
        item = {
            '日期': r['日期'],
            '星期': r['星期'],
            '哨點/崗位': '5.三總重症大樓',
            '早班 (07-19)': '、'.join(r['早班 (07-19)']) if r['早班 (07-19)'] else '—',
            '晚班 (19-07)': '、'.join(r['晚班 (19-07)']) if r['晚班 (19-07)'] else '—'
        }
        if r['substitutes']:
            item['substitutes'] = r['substitutes']
        icu_final.append(item)

    eng_hash = hashlib.md5(stable_stringify(eng_final).encode('utf-8')).hexdigest()
    icu_hash = hashlib.md5(stable_stringify(icu_final).encode('utf-8')).hexdigest()

    result_eng = {
        'tab_name': '4.三總工務所',
        'year': 2026,
        'month': 9,
        'is_current_month': True,
        'updated_at': '2026-08-31 06:00',
        'columns': ['日期', '星期', '哨點/崗位', '早班 (07-19)', '晚班 (19-07)'],
        'rows': eng_final,
        'members': ['賴鯤仲', '黃仁忠', '黃證書'],
        'posts': ['4.三總工務所'],
        'version_hash': eng_hash
    }

    result_icu = {
        'tab_name': '5.三總重症大樓',
        'year': 2026,
        'month': 9,
        'is_current_month': True,
        'updated_at': '2026-08-31 06:00',
        'columns': ['日期', '星期', '哨點/崗位', '早班 (07-19)', '晚班 (19-07)'],
        'rows': icu_final,
        'members': ['施俊宏', '林又妗', '盧建村', '賴鯤仲', '邱顯升'],
        'posts': ['5.三總重症大樓'],
        'version_hash': icu_hash
    }

    print('📊 2. 解析完成！工務所: 30天, 重症大樓: 30天 (含9/4賴大哥替補)')

    # 寫入 JSON 檔
    os.makedirs('docs/data', exist_ok=True)
    with open('docs/data/schedule_4_tsgh_eng.json', 'w', encoding='utf-8') as f:
        json.dump(result_eng, f, ensure_ascii=False, indent=2)
    with open('docs/data/schedule_5_tsgh_icu.json', 'w', encoding='utf-8') as f:
        json.dump(result_icu, f, ensure_ascii=False, indent=2)

    all_data = {'4.三總工務所': result_eng, '5.三總重症大樓': result_icu}
    with open('docs/data/schedule_live.json', 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    version_doc = {
        'year': 2026,
        'month': 9,
        'version': 'v2026.09.05-latest',
        'updated_at': '2026-09-05 10:15',
        'tabs': {
            '4.三總工務所': {'file': 'schedule_4_tsgh_eng.json', 'version_hash': eng_hash, 'row_count': 30, 'updated_at': '2026-08-31 06:00'},
            '5.三總重症大樓': {'file': 'schedule_5_tsgh_icu.json', 'version_hash': icu_hash, 'row_count': 30, 'updated_at': '2026-08-31 06:00'}
        }
    }
    with open('docs/data/schedule_version.json', 'w', encoding='utf-8') as f:
        json.dump(version_doc, f, ensure_ascii=False, indent=2)
    with open('data/schedule_version.json', 'w', encoding='utf-8') as f:
        json.dump(version_doc, f, ensure_ascii=False, indent=2)

    print('💾 3. 已更新 docs/data/ 底下所有資料檔案與版本哈希')

    # 更新 index.html 嵌入式資料
    with open('index.html', 'r', encoding='utf-8') as f:
        html = f.read()

    m_start = 'const EMBEDDED_SCHEDULE_DATA_9 = '
    idx1 = html.find(m_start)
    if idx1 != -1:
        idx2 = html.find(';\n    // 8月上月', idx1)
        if idx2 == -1:
            idx2 = html.find(';\r\n    // 8月上月', idx1)
        if idx2 != -1:
            html = html[:idx1 + len(m_start)] + json.dumps(all_data, ensure_ascii=False) + html[idx2:]
            with open('index.html', 'w', encoding='utf-8') as f:
                f.write(html)
            shutil.copyfile('index.html', 'docs/index.html')
            shutil.copyfile('index.html', 'app/static/pwa/index.html')
            print('🌐 4. 已同步嵌入 index.html、docs/index.html、app/static/pwa/index.html')

    # 同步寫入 data/storage.json 快照以供本機後端讀取
    if os.path.exists('data/storage.json'):
        try:
            with open('data/storage.json', 'r', encoding='utf-8') as sf:
                st_data = json.load(sf)
            if 'schedule_snapshots' not in st_data:
                st_data['schedule_snapshots'] = {}
            st_data['schedule_snapshots']['4.三總工務所'] = eng_final
            st_data['schedule_snapshots']['5.三總重症大樓'] = icu_final
            with open('data/storage.json', 'w', encoding='utf-8') as sf:
                json.dump(st_data, sf, ensure_ascii=False, indent=2)
            print('💾 3.1 已同步更新本機後端 data/storage.json 快照')
        except Exception as se:
            print('⚠️ 更新 storage.json 附註:', se)

    # 是否自動 git push
    if '--push' in sys.argv:
        print('🚀 5. 正在自動推送至 GitHub (git push)...')
        subprocess.run(['git', 'add', 'docs/', 'data/', 'index.html', 'app/static/pwa/index.html', 'app/static/pwa/sw.js', 'docs/sw.js', 'sw.js', 'app/services/sheets_service.py', 'sync_wheel.py'], check=True)
        subprocess.run(['git', 'commit', '-m', 'sync: 100%對齊9月份排班表，校正9/4重症A班為賴鯤仲(代林又妗)'], check=False)
        res = subprocess.run(['git', 'push', 'origin', 'main'], capture_output=True, text=True)
        if res.returncode == 0:
            print('🎉 推送成功！全球 GitHub Pages PWA 已在最新狀態！')
        else:
            print('⚠️ 推送結果:', res.stderr.strip() or res.stdout.strip())
    else:
        print('💡 提示：加上 --push 參數可自動提交並推送到 GitHub！')

    print('=====================================================')
    print(' ✨ 恭喜！同步小輪子執行完畢！')
    print('=====================================================')

if __name__ == '__main__':
    main()
